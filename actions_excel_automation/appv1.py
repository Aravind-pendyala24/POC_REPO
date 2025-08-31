from flask import Flask, render_template, request, redirect, url_for, flash
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook
import requests
import os
import json
from datetime import datetime
from filelock import FileLock

app = Flask(__name__)
load_dotenv()
app.secret_key = os.getenv("SECRET_KEY")

# GitHub configs
GITHUB_TOKEN = os.getenv("GITHUB_TOKEN")
GITHUB_OWNER = os.getenv("GITHUB_OWNER")

if not GITHUB_TOKEN or not GITHUB_OWNER:
    raise Exception("❌ Missing GITHUB_TOKEN or GITHUB_OWNER in .env file")

WORKFLOW_MAP = {
    "deployment": "deploy.yml",
    "test": "test_poc.yml"
}

# Local JSON (instead of Azure Blob for POC)
with open("apps.json") as f:
    APP_MAP = json.load(f)  # {"app1": {"repo": "frontend-app1"}, ...}

ENVIRONMENTS = ["dev", "qa", "uat", "prod"]
JOB_TYPES = list(WORKFLOW_MAP.keys())

def update_excel_local(data):
    filename = "audit_log.xlsx"
    lock = FileLock(f"{filename}.lock")

    with lock:
        try:
            wb = load_workbook(filename)
            ws = wb.active
        except Exception:
            wb = Workbook()
            ws = wb.active

        headers = ["Application", "Version", "Environment", "JobType", "Branch", "Repo", "Status", "Timestamp"]
        if ws.max_row == 0 or [cell.value for cell in ws[1]] != headers:
            ws.delete_rows(1, ws.max_row)
            ws.append(headers)

        ws.append([*data, datetime.now()])
        wb.save(filename)

def trigger_github_actions(app_name, version, environment, job_type, branch):
    repo_name = APP_MAP.get(app_name, {}).get("repo")

    if not repo_name:
        raise ValueError(f"No repo found for application '{app_name}'")

    if not app_name or not environment or not job_type:
        raise ValueError("Invalid inputs passed to the workflow.")

    if environment not in ENVIRONMENTS:
        raise ValueError(f"Eenvironment '{environment}' not found in the repo.")

    workflow_file = WORKFLOW_MAP.get(job_type)
    if not workflow_file:
        raise ValueError(f"Workflow file for job type '{job_type}' not found in the repo.")

    url = f"https://api.github.com/repos/{GITHUB_OWNER}/{repo_name}/actions/workflows/{workflow_file}/dispatches"
    headers = {
        "Authorization": f"Bearer {GITHUB_TOKEN}",
        "Accept": "application/vnd.github+json"
    }
    payload = {
        "ref": branch,
        "inputs": {
            "application": app_name,
            "version": version,
            "environment": environment
        }
    }

    print(f"branch is {branch}")
    print(f"repo_name is {repo_name}")
    print(f"payload is {payload}")
    print(f"url is {url}")

    r = requests.post(url, headers=headers, json=payload)
    if r.status_code not in (200, 201, 204):
        raise RuntimeError(f"GitHub API failed: {r.status_code}, {r.text}")

@app.route("/health")
def health_check():
    return {"status": "healthy", "timestamp": datetime.utcnow()}

@app.route("/error")
def error_page():
    msg = request.args.get("msg", "An unexpected error occurred.")
    return render_template("errorv1.html", message=msg)

# ✅ Default route: Search page
@app.route("/", methods=["GET", "POST"])
def search_page():
    answer = None
    query = None

    if request.method == "POST":
        query = request.form["query"].strip()

        # If query contains "deployment request" → redirect to deployment form
        if "deployment request" in query.lower():
            return redirect(url_for("index"))

        # Otherwise just echo back a placeholder (RAG can be integrated here)
        answer = f"📖 No deployment request found. Answer from docs (stub): You asked '{query}'"

    return render_template("qa.html", query=query, answer=answer)

@app.route("/deploy", methods=["GET", "POST"])
def index():
    if request.method == "POST":
        app_name = request.form["application"]
        version = request.form["version"]
        environment = request.form["environment"]
        job_type = request.form["job_type"]
        branch = APP_MAP[app_name].get("default_branch")

        try:
            repo_name = APP_MAP[app_name]["repo"]

            trigger_github_actions(app_name, version, environment, job_type, branch)

            success_msg = f"✅ Successfully triggered {job_type} workflow for {app_name} ({version}) in {environment}"

            update_excel_local([app_name, version, environment, job_type, branch, repo_name, success_msg])
            
            flash(success_msg, "success")

            return redirect(url_for("index"))

        except Exception as e:
            error_msg = f"❌ Failed for {app_name} ({version}) in {environment}: {str(e)}"
            repo_name = APP_MAP.get(app_name, {}).get("repo", "N/A")
            update_excel_local([app_name, version, environment, job_type, branch, repo_name, error_msg])
            return redirect(url_for("error_page", msg=str(e)))

    return render_template("form3.html",
                           applications=list(APP_MAP.keys()),
                           environments=ENVIRONMENTS,
                           job_types=JOB_TYPES,
                           applications_map=APP_MAP)

if __name__ == "__main__":
    app.run(debug=True)
