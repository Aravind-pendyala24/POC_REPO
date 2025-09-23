# app_fastapiv1.py
import os
import json
import time
import threading
import hashlib
import logging
from typing import List, Dict, Tuple
from concurrent.futures import ThreadPoolExecutor
from functools import lru_cache

import requests
from bs4 import BeautifulSoup
from dotenv import load_dotenv

from fastapi import FastAPI, Request, Form
from fastapi.responses import HTMLResponse, RedirectResponse
from fastapi.templating import Jinja2Templates

import numpy as np
import faiss
from openai import OpenAI  # Azure OpenAI client (pip package openai-compatible)

# Excel & file locking
from openpyxl import Workbook, load_workbook
from filelock import FileLock
import queue

# ---------------------------
# Logging
# ---------------------------
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("app_fastapiv1")

# ---------------------------
# Load environment
# ---------------------------
load_dotenv()

# Confluence / RAG config (if used)
CONFLUENCE_URL = os.getenv("CONFLUENCE_URL", "").rstrip("/")
ATLASSIAN_EMAIL = os.getenv("ATLASSIAN_EMAIL", "")
ATLASSIAN_TOKEN = os.getenv("ATLASSIAN_TOKEN", "")
CONFLUENCE_PAGE_IDS = [p.strip() for p in os.getenv("CONFLUENCE_PAGE_IDS", "").split(",") if p.strip()]

# Azure OpenAI config
AZURE_OPENAI_KEY = os.getenv("AZURE_OPENAI_KEY")
AZURE_OPENAI_ENDPOINT = os.getenv("AZURE_OPENAI_ENDPOINT")
AZURE_OPENAI_EMBEDDING_DEPLOYMENT = os.getenv("AZURE_OPENAI_EMBEDDING_DEPLOYMENT")
AZURE_OPENAI_CHAT_DEPLOYMENT = os.getenv("AZURE_OPENAI_CHAT_DEPLOYMENT")
AZURE_OPENAI_API_VERSION = os.getenv("AZURE_OPENAI_API_VERSION", "2024-05-01-preview")

# Jira config
JIRA_BASE_URL = os.getenv("JIRA_BASE_URL", "").rstrip("/")
JIRA_EMAIL = os.getenv("JIRA_EMAIL", "")
JIRA_API_TOKEN = os.getenv("JIRA_API_TOKEN", "")
JIRA_BOARD_ID = os.getenv("JIRA_BOARD_ID")  # optional default board id
JIRA_DEFAULT_JQL = os.getenv("JIRA_DEFAULT_JQL", "statusCategory != Done")

# Operational config
REFRESH_INTERVAL = int(os.getenv("REFRESH_INTERVAL", "300"))  # background refresh every 5m
CHUNK_SIZE = int(os.getenv("CHUNK_SIZE", "1000"))
CHUNK_OVERLAP = int(os.getenv("CHUNK_OVERLAP", "200"))

# Persistence filenames
INDEX_PATH = "faiss.index"
META_PATH = "page_metadata.json"
EXCEL_FILENAME = "audit_log.xlsx"
EXCEL_LOCKFILE = f"{EXCEL_FILENAME}.lock"

# Excel queue
excel_queue: "queue.Queue[Dict]" = queue.Queue()

# Azure OpenAI client
client = OpenAI(api_key=AZURE_OPENAI_KEY)

# FastAPI app and templates
app = FastAPI()
templates = Jinja2Templates(directory="templates")

# In-memory state
index_lock = threading.RLock()
faiss_index = None
DOCS: List[Dict] = []
PAGE_METADATA: Dict[str, Dict] = {}
cache_version = 0

# Try load metadata if present
if os.path.exists(META_PATH):
    try:
        with open(META_PATH, "r", encoding="utf-8") as f:
            PAGE_METADATA = json.load(f)
            logger.info("Loaded page metadata")
    except Exception as e:
        logger.warning("Failed to load page metadata: %s", e)
        PAGE_METADATA = {}

# ---------------------------
# Utility functions
# ---------------------------

def compute_hash(text: str) -> str:
    return hashlib.sha256(text.encode("utf-8")).hexdigest()

def chunk_text(text: str, max_chars: int = CHUNK_SIZE, overlap: int = CHUNK_OVERLAP) -> List[str]:
    if not text:
        return []
    chunks = []
    start = 0
    n = len(text)
    step = max_chars - overlap if max_chars > overlap else max_chars
    while start < n:
        end = min(start + max_chars, n)
        chunk = text[start:end].strip()
        if chunk:
            chunks.append(chunk)
        start += step
    return chunks

def fetch_confluence_page(page_id: str, timeout: int = 10) -> Dict:
    url = f"{CONFLUENCE_URL}/rest/api/content/{page_id}"
    params = {"expand": "body.storage,version"}
    auth = (ATLASSIAN_EMAIL, ATLASSIAN_TOKEN)
    resp = requests.get(url, params=params, auth=auth, timeout=timeout)
    resp.raise_for_status()
    page = resp.json()
    title = page.get("title", f"page-{page_id}")
    html = page.get("body", {}).get("storage", {}).get("value", "")
    text = BeautifulSoup(html, "html.parser").get_text(separator="\n")
    last_updated = page.get("version", {}).get("when") or None
    return {"id": page_id, "title": title, "text": text, "last_updated": last_updated}

def normalize(vectors: np.ndarray) -> np.ndarray:
    norms = np.linalg.norm(vectors, axis=1, keepdims=True)
    norms = np.where(norms == 0, 1e-12, norms)
    return vectors / norms

def embed_texts(texts: List[str]) -> np.ndarray:
    if not texts:
        return np.zeros((0, 0), dtype=np.float32)
    response = client.embeddings.create(
        model=AZURE_OPENAI_EMBEDDING_DEPLOYMENT,
        input=texts,
        extra_query={"api-version": AZURE_OPENAI_API_VERSION},
        base_url=f"{AZURE_OPENAI_ENDPOINT}/openai/deployments/{AZURE_OPENAI_EMBEDDING_DEPLOYMENT}"
    )
    vectors = [np.array(item.embedding, dtype=np.float32) for item in response.data]
    stacked = np.vstack(vectors)
    return normalize(stacked)

# ---------------------------
# FAISS utilities
# ---------------------------

def build_faiss_index(embeddings: np.ndarray) -> faiss.Index:
    if embeddings.size == 0:
        return None
    dim = embeddings.shape[1]
    idx = faiss.IndexFlatIP(dim)
    idx.add(np.ascontiguousarray(embeddings))
    return idx

def atomic_replace_index(new_index: faiss.Index, new_docs: List[Dict]):
    global faiss_index, DOCS
    with index_lock:
        faiss_index = new_index
        DOCS = new_docs
    logger.info("Replaced FAISS index; docs=%d", len(new_docs))

# ---------------------------
# Background refresh (full rebuild)
# ---------------------------

def refresh_vectorstore():
    global PAGE_METADATA, cache_version
    logger.info("Starting vectorstore refresh")
    if not CONFLUENCE_PAGE_IDS:
        logger.info("No Confluence pages configured; skipping refresh")
        return

    pages = []
    with ThreadPoolExecutor(max_workers=6) as ex:
        futures = [ex.submit(fetch_confluence_page, pid) for pid in CONFLUENCE_PAGE_IDS]
        for f in futures:
            try:
                pages.append(f.result())
            except Exception as e:
                logger.exception("Failed fetch page: %s", e)

    all_chunks = []
    chunk_docs = []
    for page in pages:
        pid = page["id"]
        text = page.get("text", "") or ""
        title = page.get("title", "")
        last_updated = page.get("last_updated")
        text_hash = compute_hash(text)

        chunks = chunk_text(text)
        PAGE_METADATA[pid] = {
            "hash": text_hash,
            "title": title,
            "last_updated": last_updated,
            "chunk_count": len(chunks),
        }

        for i, ch in enumerate(chunks):
            chunk_id = f"{pid}__chunk__{i}"
            all_chunks.append(ch)
            chunk_docs.append({
                "chunk_id": chunk_id,
                "page_id": pid,
                "title": title,
                "chunk_idx": i,
                "text": ch,
                "last_updated": last_updated,
            })

    embeddings = np.zeros((0,), dtype=np.float32)
    if all_chunks:
        embeddings = embed_texts(all_chunks)
        logger.info("Got embeddings shape %s", embeddings.shape)
    if embeddings.size:
        new_index = build_faiss_index(embeddings)
        atomic_replace_index(new_index, chunk_docs)
        try:
            with open(META_PATH, "w", encoding="utf-8") as f:
                json.dump(PAGE_METADATA, f, indent=2, ensure_ascii=False)
            faiss.write_index(new_index, INDEX_PATH)
        except Exception as e:
            logger.exception("Failed persist index/meta: %s", e)
    else:
        atomic_replace_index(None, [])

    cache_version += 1
    logger.info("Completed refresh; cache_version=%d", cache_version)

def start_background_refresh():
    def worker():
        while True:
            try:
                refresh_vectorstore()
            except Exception as e:
                logger.exception("Background refresh error: %s", e)
            time.sleep(REFRESH_INTERVAL)
    t = threading.Thread(target=worker, name="vectorstore-refresher", daemon=True)
    t.start()
    logger.info("Background refresher started (interval=%d s)", REFRESH_INTERVAL)

# ---------------------------
# Excel writer queue
# ---------------------------

def _write_row_to_excel(row: Dict):
    lock = FileLock(EXCEL_LOCKFILE, timeout=30)
    with lock:
        if not os.path.exists(EXCEL_FILENAME):
            wb = Workbook()
            ws = wb.active
            ws.append(["Application", "Version", "Environment", "JobType", "Repo", "RunURL", "Status", "Comments", "Timestamp"])
            wb.save(EXCEL_FILENAME)

        wb = load_workbook(EXCEL_FILENAME)
        ws = wb.active
        ts = time.strftime("%Y-%m-%dT%H:%M:%S%z")
        ws.append([
            row.get("application"),
            row.get("version"),
            row.get("environment"),
            row.get("job_type"),
            row.get("repo"),
            row.get("run_url"),
            row.get("status"),
            row.get("comments"),
            ts
        ])
        wb.save(EXCEL_FILENAME)
        logger.info("Excel append: %s %s", row.get("application"), row.get("status"))

def start_excel_writer():
    def writer():
        while True:
            item = excel_queue.get()
            try:
                _write_row_to_excel(item)
            except Exception as e:
                logger.exception("Excel writer error: %s", e)
            finally:
                excel_queue.task_done()
    t = threading.Thread(target=writer, name="excel-writer", daemon=True)
    t.start()
    logger.info("Excel writer started")

# ---------------------------
# Retrieval & LLM
# ---------------------------

def retrieve(query: str, k: int = 3) -> List[Dict]:
    with index_lock:
        if faiss_index is None or not DOCS:
            return []
        q_emb = embed_texts([query])
        D, I = faiss_index.search(q_emb, k)
        results = []
        for idx in I[0]:
            if idx < 0 or idx >= len(DOCS):
                continue
            results.append(DOCS[idx])
        return results

def build_prompt_from_docs(query: str, docs: List[Dict]) -> str:
    pieces = []
    for d in docs:
        title = d.get("title", "")
        text = d.get("text", "")
        pieces.append(f"Title: {title}\n{text}")
    context = "\n\n".join(pieces)
    prompt = f"""You are a helpful assistant. Use the following Confluence content to answer the question.

Context:
{context}

Question: {query}
Answer concisely and cite titles when relevant.
"""
    return prompt

def chat_completion(prompt: str, max_tokens: int = 300) -> str:
    response = client.chat.completions.create(
        model=AZURE_OPENAI_CHAT_DEPLOYMENT,
        messages=[
            {"role": "system", "content": "You are a documentation assistant."},
            {"role": "user", "content": prompt}
        ],
        max_tokens=max_tokens,
        extra_query={"api-version": AZURE_OPENAI_API_VERSION},
        base_url=f"{AZURE_OPENAI_ENDPOINT}/openai/deployments/{AZURE_OPENAI_CHAT_DEPLOYMENT}"
    )
    return response.choices[0].message.content

@lru_cache(maxsize=512)
def cached_answer(query: str, v: int) -> Tuple[str, List[Dict]]:
    try:
        docs = retrieve(query, k=3)
        prompt = build_prompt_from_docs(query, docs)
        answer = chat_completion(prompt)
        return answer, docs
    except Exception as e:
        logger.exception("cached_answer error: %s", e)
        return f"Error generating answer: {e}", []

def answer_query(query: str) -> Tuple[str, List[Dict]]:
    return cached_answer(query, cache_version)

# ---------------------------
# Jira integration
# ---------------------------

from requests.auth import HTTPBasicAuth
import re

def _jira_auth():
    if not (JIRA_EMAIL and JIRA_API_TOKEN and JIRA_BASE_URL):
        raise RuntimeError("JIRA_EMAIL, JIRA_API_TOKEN, JIRA_BASE_URL must be set")
    return HTTPBasicAuth(JIRA_EMAIL, JIRA_API_TOKEN)

def _jira_issue_url(issue_key: str) -> str:
    return f"{JIRA_BASE_URL.rstrip('/')}/browse/{issue_key}"

def fetch_jira_board_open_issues(board_id: int = None, jql: str = None, max_results: int = 200) -> List[Dict]:
    if board_id is None:
        if not JIRA_BOARD_ID:
            raise RuntimeError("Board ID not provided and JIRA_BOARD_ID not set")
        board_id = int(JIRA_BOARD_ID)

    if jql is None:
        jql = JIRA_DEFAULT_JQL

    auth = _jira_auth()
    headers = {"Accept": "application/json"}
    issues = []
    start_at = 0
    page_size = 50

    while True:
        params = {
            "jql": jql,
            "startAt": start_at,
            "maxResults": page_size,
            "fields": "summary,status,assignee,labels,priority"
        }
        url = f"{JIRA_BASE_URL.rstrip('/')}/rest/agile/1.0/board/{board_id}/issue"
        resp = requests.get(url, params=params, headers=headers, auth=auth, timeout=15)
        if resp.status_code == 404:
            raise RuntimeError(f"Board {board_id} not found (404).")
        resp.raise_for_status()
        payload = resp.json()
        page_issues = payload.get("issues", [])
        for i in page_issues:
            key = i.get("key")
            fields = i.get("fields", {})
            summary = fields.get("summary", "")
            status = (fields.get("status") or {}).get("name", "")
            assignee = fields.get("assignee") or {}
            assignee_name = assignee.get("displayName") if assignee else None
            issues.append({
                "key": key,
                "summary": summary,
                "status": status,
                "assignee": assignee_name,
                "url": _jira_issue_url(key)
            })
            if len(issues) >= max_results:
                return issues

        total = payload.get("total", 0)
        start_at += len(page_issues)
        if start_at >= total or not page_issues:
            break

    return issues

JIRA_LIST_REGEX = re.compile(r"\b(list|show|display)\b.*\b(open|pending|to[-\s]*do)\b.*\b(ticket|issue|tickets|issues)\b", re.IGNORECASE)

def handle_jira_query_if_any(query: str) -> List[Dict] or None:
    if not JIRA_BASE_URL:
        return None
    if not JIRA_LIST_REGEX.search(query):
        return None

    board_id = None
    m = re.search(r"board\s+(\d+)", query, re.IGNORECASE)
    if m:
        board_id = int(m.group(1))

    try:
        issues = fetch_jira_board_open_issues(board_id=board_id)
        return issues
    except Exception as e:
        logger.exception("Error fetching Jira issues: %s", e)
        # return a single-item list with the error to render nicely
        return [{"key": "ERROR", "summary": str(e), "status": "ERROR", "assignee": None, "url": ""}]

# ---------------------------
# FastAPI routes
# ---------------------------

@app.on_event("startup")
def on_startup():
    start_background_refresh()
    start_excel_writer()

@app.get("/", response_class=HTMLResponse)
async def home(request: Request):
    return templates.TemplateResponse("qa.html", {"request": request})

@app.post("/answer", response_class=HTMLResponse)
async def answer(request: Request, query: str = Form(...)):
    # First check Jira intent
    jira_issues = None
    try:
        jira_issues = handle_jira_query_if_any(query)
    except Exception as e:
        logger.exception("Jira intent check error: %s", e)
        jira_issues = [{"key": "ERROR", "summary": f"Jira check failed: {e}", "status": "ERROR", "assignee": None, "url": ""}]

    if jira_issues is not None:
        # Render page with jira_issues table
        return templates.TemplateResponse("qa.html", {
            "request": request,
            "query": query,
            "jira_issues": jira_issues
        })

    # Otherwise use RAG/LLM to answer
    try:
        # run blocking answer_query off the main thread
        import asyncio
        answer_text, docs = await asyncio.to_thread(answer_query, query)
    except Exception as e:
        logger.exception("Answer generation error: %s", e)
        answer_text, docs = f"Error generating answer: {e}", []

    return templates.TemplateResponse("qa.html", {
        "request": request,
        "query": query,
        "answer": answer_text,
        "results": docs
    })

@app.post("/refresh_jira", response_class=HTMLResponse)
async def refresh_jira(request: Request, board_id: int = Form(None)):
    """
    Manual refresh/lookup of Jira board. Returns the same template populated with results.
    """
    try:
        issues = await app.loop.run_in_executor(None, fetch_jira_board_open_issues, int(board_id) if board_id else None)
    except Exception as e:
        logger.exception("Manual Jira refresh failed: %s", e)
        issues = [{"key": "ERROR", "summary": str(e), "status": "ERROR", "assignee": None, "url": ""}]

    return templates.TemplateResponse("qa.html", {
        "request": request,
        "query": f"Manual refresh board {board_id}" if board_id else "Manual Jira refresh",
        "jira_issues": issues
    })

@app.get("/health")
def health():
    return {"status": "ok", "index_docs": len(DOCS), "cache_version": cache_version}
