# app_fastapiv1.py
import os
import json
import time
import math
import threading
import hashlib
import logging
from typing import List, Dict, Tuple
from concurrent.futures import ThreadPoolExecutor

import requests
from bs4 import BeautifulSoup
from dotenv import load_dotenv

from fastapi import FastAPI, Request, Form
from fastapi.responses import HTMLResponse, RedirectResponse
from fastapi.templating import Jinja2Templates

import numpy as np
import faiss

from openai import OpenAI  # Azure OpenAI client instantiation pattern used earlier
from functools import lru_cache

# Excel & file locking
from openpyxl import Workbook, load_workbook
from filelock import FileLock
import queue

# ---------------------------
# Logging
# ---------------------------
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("app_fastapiv1")

# ---------------------------
# Load environment
# ---------------------------
load_dotenv()

CONFLUENCE_URL = os.getenv("CONFLUENCE_URL", "").rstrip("/")
ATLASSIAN_EMAIL = os.getenv("ATLASSIAN_EMAIL", "")
ATLASSIAN_TOKEN = os.getenv("ATLASSIAN_TOKEN", "")
CONFLUENCE_PAGE_IDS = [p.strip() for p in os.getenv("CONFLUENCE_PAGE_IDS", "").split(",") if p.strip()]

AZURE_OPENAI_KEY = os.getenv("AZURE_OPENAI_KEY")
AZURE_OPENAI_ENDPOINT = os.getenv("AZURE_OPENAI_ENDPOINT")
AZURE_OPENAI_EMBEDDING_DEPLOYMENT = os.getenv("AZURE_OPENAI_EMBEDDING_DEPLOYMENT")
AZURE_OPENAI_CHAT_DEPLOYMENT = os.getenv("AZURE_OPENAI_CHAT_DEPLOYMENT")
AZURE_OPENAI_API_VERSION = os.getenv("AZURE_OPENAI_API_VERSION", "2024-05-01-preview")

# Refresh interval (seconds)
REFRESH_INTERVAL = int(os.getenv("REFRESH_INTERVAL", "300"))  # default 5 minutes

# Chunking config
CHUNK_SIZE = int(os.getenv("CHUNK_SIZE", "1000"))      # characters per chunk
CHUNK_OVERLAP = int(os.getenv("CHUNK_OVERLAP", "200")) # overlap in chars

# Persistence
INDEX_PATH = "faiss.index"
META_PATH = "page_metadata.json"
EXCEL_FILENAME = "audit_log.xlsx"
EXCEL_LOCKFILE = f"{EXCEL_FILENAME}.lock"

# Excel queue
excel_queue: "queue.Queue[Dict]" = queue.Queue()

# Azure OpenAI client
client = OpenAI(api_key=AZURE_OPENAI_KEY)

# FastAPI + templates
app = FastAPI()
templates = Jinja2Templates(directory="templates")

# In-memory state (thread-safe guarded)
index_lock = threading.RLock()
faiss_index = None       # faiss.Index object
DOCS: List[Dict] = []    # metadata for each chunk in the index (order matches FAISS)
PAGE_METADATA: Dict[str, Dict] = {}  # page_id -> {hash, title, last_updated, chunk_count}
cache_version = 0        # increment to invalidate LRU cache

# If persisted metadata exists, try to load it (we won't persist embeddings)
if os.path.exists(META_PATH):
    try:
        with open(META_PATH, "r", encoding="utf-8") as f:
            PAGE_METADATA = json.load(f)
            logger.info("Loaded page metadata from %s", META_PATH)
    except Exception as e:
        logger.warning("Failed to load metadata: %s", e)
        PAGE_METADATA = {}

# ---------------------------
# Utility functions
# ---------------------------

def compute_hash(text: str) -> str:
    return hashlib.sha256(text.encode("utf-8")).hexdigest()

def chunk_text(text: str, max_chars: int = CHUNK_SIZE, overlap: int = CHUNK_OVERLAP) -> List[str]:
    """Simple character-based chunking with overlap."""
    if not text:
        return []
    chunks = []
    start = 0
    n = len(text)
    step = max_chars - overlap if max_chars > overlap else max_chars
    while start < n:
        end = min(start + max_chars, n)
        chunk = text[start:end].strip()
        if chunk:
            chunks.append(chunk)
        start += step
    return chunks

def fetch_confluence_page(page_id: str, timeout: int = 10) -> Dict:
    """
    Fetch Confluence page by REST API and return minimal metadata:
    { "id", "title", "text", "last_updated" }
    """
    url = f"{CONFLUENCE_URL}/rest/api/content/{page_id}"
    params = {"expand": "body.storage,version"}
    auth = (ATLASSIAN_EMAIL, ATLASSIAN_TOKEN)
    logger.debug("Fetching Confluence page %s", page_id)
    resp = requests.get(url, params=params, auth=auth, timeout=timeout)
    resp.raise_for_status()
    page = resp.json()

    title = page.get("title", f"page-{page_id}")
    html = page.get("body", {}).get("storage", {}).get("value", "")
    text = BeautifulSoup(html, "html.parser").get_text(separator="\n")
    # version.when may be missing sometimes; fallback to current time
    last_updated = page.get("version", {}).get("when") or page.get("history", {}).get("lastUpdated", {}).get("when") or None

    return {"id": page_id, "title": title, "text": text, "last_updated": last_updated}

def normalize(vectors: np.ndarray) -> np.ndarray:
    norms = np.linalg.norm(vectors, axis=1, keepdims=True)
    norms = np.where(norms == 0, 1e-12, norms)
    return vectors / norms

def embed_texts(texts: List[str]) -> np.ndarray:
    """
    Batch embedding call to Azure OpenAI.
    Returns normalized np.ndarray shape (n, dim).
    """
    if not texts:
        return np.zeros((0, 0), dtype=np.float32)
    logger.info("Requesting embeddings for batch size=%d", len(texts))
    # Use Azure-compatible call: base_url + extra_query pattern
    response = client.embeddings.create(
        model=AZURE_OPENAI_EMBEDDING_DEPLOYMENT,
        input=texts,
        extra_query={"api-version": AZURE_OPENAI_API_VERSION},
        base_url=f"{AZURE_OPENAI_ENDPOINT}/openai/deployments/{AZURE_OPENAI_EMBEDDING_DEPLOYMENT}"
    )
    vectors = [np.array(item.embedding, dtype=np.float32) for item in response.data]
    stacked = np.vstack(vectors)
    return normalize(stacked)

# ---------------------------
# FAISS / index management
# ---------------------------

def build_faiss_index(embeddings: np.ndarray) -> faiss.Index:
    """
    Build a new FAISS index for the given embeddings (already normalized).
    We'll use IndexFlatIP (inner product) which works with normalized vectors as cosine sim.
    """
    if embeddings.size == 0:
        return None
    dim = embeddings.shape[1]
    index = faiss.IndexFlatIP(dim)
    index.add(np.ascontiguousarray(embeddings))
    return index

def atomic_replace_index(new_index: faiss.Index, new_docs: List[Dict]):
    """
    Replace global FAISS index and DOCS atomically under lock.
    """
    global faiss_index, DOCS
    with index_lock:
        faiss_index = new_index
        DOCS = new_docs
    logger.info("Replaced FAISS index: %d docs", len(new_docs))

# ---------------------------
# Refresh vector store (background)
# ---------------------------

def refresh_vectorstore():
    """
    Fetch all configured Confluence pages concurrently, chunk them,
    create embeddings in batch, and rebuild FAISS index atomically.
    This is intentionally a full rebuild for correctness and simplicity.
    """
    global PAGE_METADATA, cache_version

    logger.info("Starting refresh_vectorstore")
    # 1) fetch pages in parallel
    pages: List[Dict] = []
    if not CONFLUENCE_PAGE_IDS:
        logger.warning("No Confluence page IDs configured.")
        return

    with ThreadPoolExecutor(max_workers=6) as ex:
        futures = [ex.submit(fetch_confluence_page, pid) for pid in CONFLUENCE_PAGE_IDS]
        for f in futures:
            try:
                page = f.result()
                pages.append(page)
            except Exception as e:
                logger.exception("Failed to fetch page: %s", e)

    # 2) chunk pages & prepare chunk metadata list
    all_chunks: List[str] = []
    chunk_docs: List[Dict] = []
    for page in pages:
        pid = page["id"]
        text = page.get("text", "") or ""
        title = page.get("title", "")
        last_updated = page.get("last_updated")
        text_hash = compute_hash(text)

        chunks = chunk_text(text)
        # store page metadata (hash + title + last_updated + chunk_count)
        PAGE_METADATA[pid] = {
            "hash": text_hash,
            "title": title,
            "last_updated": last_updated,
            "chunk_count": len(chunks),
        }

        for i, ch in enumerate(chunks):
            chunk_id = f"{pid}__chunk__{i}"
            all_chunks.append(ch)
            chunk_docs.append({
                "chunk_id": chunk_id,
                "page_id": pid,
                "title": title,
                "chunk_idx": i,
                "text": ch,
                "last_updated": last_updated,
            })

    # 3) embed all chunks in batch (if any)
    embeddings = np.zeros((0,), dtype=np.float32)
    if all_chunks:
        emb = embed_texts(all_chunks)  # shape (n, dim)
        embeddings = emb
        logger.info("Embeddings shape: %s", str(embeddings.shape))
    else:
        logger.info("No chunks found to embed.")

    # 4) build new FAISS index and replace atomically
    if embeddings.size:
        new_index = build_faiss_index(embeddings)
        atomic_replace_index(new_index, chunk_docs)
        # persist metadata
        try:
            with open(META_PATH, "w", encoding="utf-8") as f:
                json.dump(PAGE_METADATA, f, indent=2, ensure_ascii=False)
            # persist index file (optional)
            faiss.write_index(new_index, INDEX_PATH)
        except Exception as e:
            logger.exception("Failed to persist metadata/index: %s", e)
    else:
        # if no embeddings, clear index
        atomic_replace_index(None, [])

    # 5) bump cache version to invalidate LRU caches (answers)
    cache_version += 1
    logger.info("Finished refresh_vectorstore; cache_version=%d", cache_version)

# ---------------------------
# Background tasks
# ---------------------------

def start_background_refresh():
    """
    Start a daemon thread that periodically refreshes the vector store.
    """
    def worker():
        while True:
            try:
                refresh_vectorstore()
            except Exception as e:
                logger.exception("Error in refresh_vectorstore: %s", e)
            time.sleep(REFRESH_INTERVAL)
    t = threading.Thread(target=worker, name="vectorstore-refresher", daemon=True)
    t.start()
    logger.info("Started background refresher thread (interval=%d s)", REFRESH_INTERVAL)


def start_excel_writer():
    """
    Background thread consumes excel_queue and writes rows to the excel file with a file lock.
    Each queue item is a dict with keys matching columns.
    """
    def writer():
        while True:
            item = excel_queue.get()
            try:
                _write_row_to_excel(item)
            except Exception as e:
                logger.exception("Excel write failed: %s", e)
            finally:
                excel_queue.task_done()
    t = threading.Thread(target=writer, name="excel-writer", daemon=True)
    t.start()
    logger.info("Started excel writer thread")

def _write_row_to_excel(row: Dict):
    """
    Thread-safe Excel append with FileLock. Column order:
    Application | Version | Environment | JobType | Repo | RunURL | Status | Comments | Timestamp
    """
    lock = FileLock(EXCEL_LOCKFILE, timeout=30)
    with lock:
        if not os.path.exists(EXCEL_FILENAME):
            wb = Workbook()
            ws = wb.active
            ws.append(["Application", "Version", "Environment", "JobType", "Repo", "RunURL", "Status", "Comments", "Timestamp"])
            wb.save(EXCEL_FILENAME)

        wb = load_workbook(EXCEL_FILENAME)
        ws = wb.active
        ts = time.strftime("%Y-%m-%dT%H:%M:%S%z")
        ws.append([
            row.get("application"),
            row.get("version"),
            row.get("environment"),
            row.get("job_type"),
            row.get("repo"),
            row.get("run_url"),
            row.get("status"),
            row.get("comments"),
            ts
        ])
        wb.save(EXCEL_FILENAME)
        logger.info("Appended row to excel: %s %s", row.get("application"), row.get("status"))

# ---------------------------
# Retrieval & LLM
# ---------------------------

def retrieve(query: str, k: int = 3) -> List[Dict]:
    """
    Retrieve top-k chunk docs for the query using the current FAISS index.
    Returns list of chunk_docs (as stored in DOCS).
    """
    with index_lock:
        if faiss_index is None or not DOCS:
            return []
        # embed query
        q_emb = embed_texts([query])  # (1, dim)
        D, I = faiss_index.search(q_emb, k)
        results = []
        for idx in I[0]:
            if idx < 0 or idx >= len(DOCS):
                continue
            results.append(DOCS[idx])
        return results

def build_prompt_from_docs(query: str, docs: List[Dict]) -> str:
    pieces = []
    for d in docs:
        title = d.get("title", "")
        text = d.get("text", "")
        pieces.append(f"Title: {title}\n{text}")
    context = "\n\n".join(pieces)
    prompt = f"""You are a helpful assistant. Use the following Confluence content to answer the question.
Context:
{context}

Question: {query}
Answer concisely and cite titles when relevant.
"""
    return prompt

def chat_completion(prompt: str, max_tokens: int = 300) -> str:
    """
    Call Azure OpenAI chat/completions with the prompt and return the assistant text.
    """
    response = client.chat.completions.create(
        model=AZURE_OPENAI_CHAT_DEPLOYMENT,
        messages=[
            {"role": "system", "content": "You are a documentation assistant."},
            {"role": "user", "content": prompt}
        ],
        max_tokens=max_tokens,
        extra_query={"api-version": AZURE_OPENAI_API_VERSION},
        base_url=f"{AZURE_OPENAI_ENDPOINT}/openai/deployments/{AZURE_OPENAI_CHAT_DEPLOYMENT}"
    )
    return response.choices[0].message.content

# LRU-cache wrapper — include cache_version so that cache is invalidated after refresh
@lru_cache(maxsize=512)
def cached_answer(query: str, v: int) -> Tuple[str, List[Dict]]:
    """
    query: user query string
    v: cache_version integer (include to invalidate when index refreshes)
    Returns (answer_str, results_list)
    """
    try:
        # retrieve top docs using fastest path
        docs = retrieve(query, k=3)
        prompt = build_prompt_from_docs(query, docs)
        answer = chat_completion(prompt)
        return answer, docs
    except Exception as e:
        logger.exception("Error in cached_answer: %s", e)
        return f"Error generating answer: {e}", []

def answer_query(query: str) -> Tuple[str, List[Dict]]:
    """
    Public wrapper to call the cached answer with current cache_version.
    """
    return cached_answer(query, cache_version)

# ---------------------------
# FastAPI routes
# ---------------------------

@app.on_event("startup")
def on_startup():
    # start background tasks
    start_background_refresh()
    start_excel_writer()

@app.get("/", response_class=HTMLResponse)
async def home(request: Request):
    return templates.TemplateResponse("qa.html", {"request": request})

@app.post("/answer", response_class=HTMLResponse)
async def answer(request: Request, query: str = Form(...)):
    """
    Handles the query submission. Uses the cached answer (fast).
    Renders a result page (qa.html) with the answer.
    """
    # do the expensive work off main thread to avoid blocking (though cached_answer may be fast)
    try:
        answer_text, docs = await app.loop.run_in_executor(None, answer_query, query)
    except Exception as e:
        logger.exception("Error while answering: %s", e)
        answer_text, docs = f"Error: {e}", []

    return templates.TemplateResponse("qa.html", {
        "request": request,
        "query": query,
        "answer": answer_text,
        "results": docs
    })

# Optional health endpoint
@app.get("/health")
def health():
    return {"status": "ok", "index_docs": len(DOCS), "cache_version": cache_version}



# Notes, reasoning & next steps

# Why a full rebuild?
# Rebuilding the index each refresh is simpler and avoids tricky bookkeeping to remove/replace old chunks. If you have many pages and rebuild cost becomes unacceptable, we can change to incremental embedding by storing embeddings per chunk and using FAISS IndexIDMap to add/remove IDs.

# Batch embedding
# embed_texts() sends a list of texts to the Azure embeddings endpoint at once — this is much faster than one-by-one calls.

# Chunking
# Character-based chunking with overlap ensures long pages are broken into searchable pieces. You can refine chunk size to balance precision and cost.

# LRU cache invalidation
# cache_version is incremented on refresh. The cached_answer function includes v in its signature, so the LRU cache is automatically invalidated after refresh.

# Excel writes
# Writes are queued and executed by a background thread so user requests are fast and not blocked by disk I/O.

# Thread-safety
# index_lock protects replacing the FAISS index and reading it during retrieval.

# Persistent files
# page_metadata.json and faiss.index are written after each successful refresh. You can remove index persistence if your environment requires (or switch to a vector DB).

# Performance

# The user-facing request path no longer rebuilds the index; it uses the latest background-built index.

# Embedding and rebuilding still happen periodically; tune REFRESH_INTERVAL to your needs.
